import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from ofxparse import OfxParser
from openpyxl import load_workbook

from app.models.receivable import Receivable
from app.models.payment import Payment

# Patterns for column detection
CNPJ_PATTERN = re.compile(r"^\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}$")
CPF_PATTERN = re.compile(r"^\d{3}\.?\d{3}\.?\d{3}-?\d{2}$")
DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y"]

# Column name heuristics (Portuguese)
VALUE_HINTS = ["valor", "value", "montante", "amount", "total", "face_value", "vl_", "vlr"]
DATE_HINTS = ["data", "date", "vencimento", "due", "dt_", "emissao", "pagamento"]
CNPJ_HINTS = ["cnpj", "cpf", "documento", "doc", "sacado", "pagador", "cedente", "devedor"]
NAME_HINTS = ["nome", "name", "razao", "razão", "sacado", "pagador", "cedente", "devedor"]


def normalize_cnpj(value: str) -> str:
    """Remove punctuation from CNPJ/CPF."""
    return re.sub(r"[.\-/]", "", value.strip())


def parse_monetary_value(value: str) -> Decimal | None:
    """Parse Brazilian monetary values: 1.234,56 or 1234.56"""
    if not value or not value.strip():
        return None
    v = value.strip().replace("R$", "").replace(" ", "")
    # Brazilian format: 1.234,56
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    elif "," in v:
        v = v.replace(",", ".")
    try:
        return Decimal(v)
    except InvalidOperation:
        return None


def parse_date(value: str) -> date | None:
    """Try multiple date formats."""
    if not value or not value.strip():
        return None
    v = value.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def detect_column_type(header: str, sample_values: list[str]) -> str:
    """Detect what type of data a column contains."""
    h = header.lower().strip()

    # Check header name first — more specific hints take priority
    # "nome" and "name" are checked BEFORE generic hints like "sacado"
    if any(hint in h for hint in ["nome", "name", "razao", "razão"]):
        return "name"
    if any(hint in h for hint in ["cnpj", "cpf", "documento", "doc"]):
        return "cnpj"
    if any(hint in h for hint in VALUE_HINTS):
        return "value"
    if any(hint in h for hint in DATE_HINTS):
        return "date"
    # Generic hints (could be name or cnpj — use sample values to decide)
    if any(hint in h for hint in ["sacado", "pagador", "cedente", "devedor"]):
        non_empty = [v for v in sample_values if v and v.strip()]
        cnpj_matches = sum(1 for v in non_empty if CNPJ_PATTERN.match(v.strip()) or CPF_PATTERN.match(v.strip()))
        if cnpj_matches > len(non_empty) * 0.3:
            return "cnpj"
        return "name"

    # Check sample values
    non_empty = [v for v in sample_values if v and v.strip()]
    if not non_empty:
        return "unknown"

    cnpj_matches = sum(1 for v in non_empty if CNPJ_PATTERN.match(v.strip()) or CPF_PATTERN.match(v.strip()))
    if cnpj_matches > len(non_empty) * 0.5:
        return "cnpj"

    value_matches = sum(1 for v in non_empty if parse_monetary_value(v) is not None)
    if value_matches > len(non_empty) * 0.7:
        return "value"

    date_matches = sum(1 for v in non_empty if parse_date(v) is not None)
    if date_matches > len(non_empty) * 0.7:
        return "date"

    return "name"


def parse_csv_content(content: bytes) -> dict:
    """Parse CSV content with smart column detection."""
    # Try different encodings
    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode file. Try UTF-8 or Latin-1 encoding.")

    reader = csv.reader(io.StringIO(text), delimiter=_detect_delimiter(text))
    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("File must have at least a header row and one data row.")

    headers = rows[0]
    data_rows = rows[1:]

    # Detect column types
    column_map: dict[str, list[int]] = {}
    for i, header in enumerate(headers):
        samples = [row[i] for row in data_rows[:20] if i < len(row)]
        col_type = detect_column_type(header, samples)
        if col_type not in column_map:
            column_map[col_type] = []
        column_map[col_type].append(i)

    file_type = _detect_file_type(headers)
    return _build_records(column_map, data_rows, "csv", file_type)


def parse_xlsx_content(content: bytes) -> dict:
    """Parse XLSX content with smart column detection."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])

    if len(rows) < 2:
        raise ValueError("File must have at least a header row and one data row.")

    headers = rows[0]
    data_rows = rows[1:]

    column_map = {}
    for i, header in enumerate(headers):
        samples = [row[i] for row in data_rows[:20] if i < len(row)]
        col_type = detect_column_type(header, samples)
        if col_type not in column_map:
            column_map[col_type] = []
        column_map[col_type].append(i)

    file_type = _detect_file_type(headers)
    return _build_records(column_map, data_rows, "xlsx", file_type)


def _detect_delimiter(text: str) -> str:
    """Detect CSV delimiter (comma, semicolon, tab)."""
    first_line = text.split("\n")[0]
    for delim in [";", ",", "\t", "|"]:
        if delim in first_line:
            return delim
    return ","


def _detect_file_type(headers: list[str]) -> str:
    """Detect if file contains receivables or payments based on headers."""
    all_headers = " ".join(h.lower() for h in headers)
    payment_hints = ["pagamento", "pagador", "pago", "paga", "payment", "paid", "extrato"]
    receivable_hints = ["recebivel", "recebiveis", "sacado", "cedente", "vencimento", "receivable"]

    pay_score = sum(1 for h in payment_hints if h in all_headers)
    rec_score = sum(1 for h in receivable_hints if h in all_headers)

    if pay_score > rec_score:
        return "payment"
    return "receivable"


def _build_records(column_map: dict, data_rows: list, source: str, file_type: str = "receivable") -> dict:
    """Build Receivable/Payment records from detected columns."""
    receivables = []
    payments = []
    errors = []

    value_cols = column_map.get("value", [])
    date_cols = column_map.get("date", [])
    cnpj_cols = column_map.get("cnpj", [])
    name_cols = column_map.get("name", [])

    if not value_cols:
        raise ValueError("Could not detect a value/amount column in the file.")

    for row_idx, row in enumerate(data_rows):
        try:
            value = None
            for col in value_cols:
                if col < len(row):
                    value = parse_monetary_value(row[col])
                    if value is not None:
                        break

            if value is None or value == 0:
                continue

            record_date = None
            for col in date_cols:
                if col < len(row):
                    record_date = parse_date(row[col])
                    if record_date is not None:
                        break

            cnpj = None
            for col in cnpj_cols:
                if col < len(row):
                    raw = row[col].strip()
                    if raw:
                        cnpj = normalize_cnpj(raw)
                        break

            name = None
            for col in name_cols:
                if col < len(row):
                    raw = row[col].strip()
                    if raw:
                        name = raw
                        break

            abs_value = abs(value)
            is_payment = file_type == "payment" or value < 0

            if is_payment:
                payments.append(Payment(
                    payer_cnpj=cnpj,
                    payer_name=name,
                    amount=abs_value,
                    date=record_date,
                    source=source,
                ))
            else:
                receivables.append(Receivable(
                    debtor_cnpj=cnpj,
                    debtor_name=name,
                    face_value=abs_value,
                    due_date=record_date,
                    source=source,
                ))

        except Exception as e:
            errors.append({"row": row_idx + 2, "error": str(e)})

    return {"receivables": receivables, "payments": payments, "errors": errors}


def parse_ofx_content(content: bytes) -> dict:
    """Parse OFX bank statement. All transactions become payments."""
    try:
        ofx = OfxParser.parse(io.BytesIO(content))
    except Exception:
        raise ValueError("Could not parse OFX file. Make sure it is a valid bank statement.")

    payments = []
    errors = []

    if not ofx.account or not ofx.account.statement or not ofx.account.statement.transactions:
        raise ValueError("OFX file has no transactions.")

    for i, txn in enumerate(ofx.account.statement.transactions):
        try:
            amount = Decimal(str(txn.amount))
            if amount == 0:
                continue

            txn_date = txn.date.date() if isinstance(txn.date, datetime) else txn.date

            # OFX memo/payee often contain the payer name
            payer_name = txn.payee or txn.memo or None
            if payer_name:
                payer_name = payer_name.strip()

            # Try to extract CNPJ/CPF from memo or check number
            payer_cnpj = None
            search_text = f"{txn.memo or ''} {txn.payee or ''} {txn.checknum or ''}"
            # Use patterns without anchors for searching within text
            cnpj_search = re.search(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}", search_text)
            cpf_search = re.search(r"\d{3}\.?\d{3}\.?\d{3}-?\d{2}", search_text)
            if cnpj_search:
                payer_cnpj = normalize_cnpj(cnpj_search.group())
            elif cpf_search:
                payer_cnpj = normalize_cnpj(cpf_search.group())

            payments.append(Payment(
                payer_cnpj=payer_cnpj,
                payer_name=payer_name,
                amount=abs(amount),
                date=txn_date,
                bank_reference=txn.id or txn.checknum or None,
                source="ofx",
            ))
        except Exception as e:
            errors.append({"row": i + 1, "error": str(e)})

    return {"receivables": [], "payments": payments, "errors": errors}


# ---------------------------------------------------------------------------
# CNAB 240 / 400 parser (retorno de cobrança)
# ---------------------------------------------------------------------------

# Occurrence codes that mean "paid/settled"
CNAB_PAID_CODES = {"06", "07", "10", "17"}  # liquidação, parcial, baixa+liq, liq após baixa


def _detect_cnab_format(content: bytes) -> str | None:
    """Detect CNAB format by line length. Returns '240', '400', or None."""
    for encoding in ["latin-1", "cp1252", "utf-8"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return None

    lines = [l for l in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if l.strip()]
    if not lines:
        return None

    length = len(lines[0])
    if length == 240:
        return "240"
    elif length == 400:
        return "400"
    return None


def _cnab_parse_value(raw: str, decimals: int = 2) -> Decimal:
    """Parse CNAB fixed-width numeric value (e.g. '000001234567' → 12345.67)."""
    raw = raw.strip()
    if not raw or not raw.isdigit():
        return Decimal("0")
    integer_part = raw[:-decimals] if decimals else raw
    decimal_part = raw[-decimals:] if decimals else ""
    return Decimal(f"{int(integer_part)}.{decimal_part}")


def _cnab_parse_date(raw: str) -> date | None:
    """Parse CNAB date in DDMMYYYY (8 chars) or DDMMYY (6 chars)."""
    raw = raw.strip()
    if not raw or raw == "00000000" or raw == "000000" or not raw.isdigit():
        return None
    try:
        if len(raw) == 8:
            return datetime.strptime(raw, "%d%m%Y").date()
        elif len(raw) == 6:
            return datetime.strptime(raw, "%d%m%y").date()
    except ValueError:
        pass
    return None


def _cnab_extract_cnpj(raw: str) -> str | None:
    """Extract and normalize CNPJ/CPF from CNAB fixed-width field."""
    raw = raw.strip()
    if not raw or raw == "0" * len(raw):
        return None
    # Remove leading zeros for CPF (11 digits) or CNPJ (14 digits)
    digits = raw.lstrip("0")
    if not digits:
        return None
    # Pad back to 11 or 14
    full = raw.strip()
    if len(full) <= 11:
        return full.zfill(11)
    return full.zfill(14)


def parse_cnab240_content(content: bytes) -> dict:
    """Parse CNAB 240 retorno file. Segments T+U → Payment records."""
    for encoding in ["latin-1", "cp1252", "utf-8"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Não foi possível decodificar o arquivo CNAB.")

    lines = [l for l in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if l.strip()]

    receivables = []
    payments = []
    errors = []

    # Parse segments T and U in pairs
    seg_t = None
    for line_num, line in enumerate(lines, 1):
        if len(line) < 240:
            continue

        tipo_registro = line[7]  # position 8 (0-indexed: 7)

        if tipo_registro != "3":  # only detail records
            continue

        segmento = line[13].upper()  # position 14

        if segmento == "T":
            try:
                codigo_ocorrencia = line[15:17].strip()
                nosso_numero = line[40:48].strip()
                numero_documento = line[58:68].strip()
                vencimento = _cnab_parse_date(line[73:81])
                valor_titulo = _cnab_parse_value(line[81:96])
                tipo_inscricao = line[132]  # 1=CPF, 2=CNPJ
                inscricao_pagador = line[133:148]
                nome_pagador = line[148:178].strip()

                payer_cnpj = _cnab_extract_cnpj(inscricao_pagador)

                seg_t = {
                    "ocorrencia": codigo_ocorrencia,
                    "nosso_numero": nosso_numero,
                    "numero_documento": numero_documento,
                    "vencimento": vencimento,
                    "valor_titulo": valor_titulo,
                    "payer_cnpj": payer_cnpj,
                    "payer_name": nome_pagador if nome_pagador else None,
                    "line": line_num,
                }
            except Exception as e:
                errors.append({"row": line_num, "error": str(e)})
                seg_t = None

        elif segmento == "U" and seg_t is not None:
            try:
                valor_pago = _cnab_parse_value(line[77:92])
                data_ocorrencia = _cnab_parse_date(line[137:145])
                data_credito = _cnab_parse_date(line[145:153])
                juros_multa = _cnab_parse_value(line[17:32])
                desconto = _cnab_parse_value(line[32:47])

                is_paid = seg_t["ocorrencia"] in CNAB_PAID_CODES

                # Receivable: always create from seg T (the boleto)
                receivables.append(Receivable(
                    debtor_cnpj=seg_t["payer_cnpj"],
                    debtor_name=seg_t["payer_name"],
                    face_value=seg_t["valor_titulo"],
                    due_date=seg_t["vencimento"],
                    status="conciliated" if is_paid else "pending",
                    source="cnab240",
                ))

                # Payment: only if paid
                if is_paid and valor_pago > 0:
                    ref = seg_t["nosso_numero"] or seg_t["numero_documento"]
                    payments.append(Payment(
                        payer_cnpj=seg_t["payer_cnpj"],
                        payer_name=seg_t["payer_name"],
                        amount=valor_pago,
                        date=data_credito or data_ocorrencia,
                        bank_reference=ref if ref else None,
                        source="cnab240",
                    ))
            except Exception as e:
                errors.append({"row": line_num, "error": str(e)})
            finally:
                seg_t = None

    return {"receivables": receivables, "payments": payments, "errors": errors}


def parse_cnab400_content(content: bytes) -> dict:
    """Parse CNAB 400 retorno file. Detail records → Payment records."""
    for encoding in ["latin-1", "cp1252", "utf-8"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Não foi possível decodificar o arquivo CNAB.")

    lines = [l for l in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if l.strip()]

    receivables = []
    payments = []
    errors = []

    for line_num, line in enumerate(lines, 1):
        if len(line) < 400:
            continue

        tipo_registro = line[0]
        if tipo_registro != "1":  # only detail records
            continue

        try:
            codigo_ocorrencia = line[108:110].strip()
            data_ocorrencia = _cnab_parse_date(line[110:116])
            numero_documento = line[116:126].strip()
            nosso_numero = line[62:70].strip()
            vencimento = _cnab_parse_date(line[146:152])
            valor_titulo = _cnab_parse_value(line[152:165])
            valor_pago = _cnab_parse_value(line[253:266])
            juros_multa = _cnab_parse_value(line[266:279])
            desconto = _cnab_parse_value(line[240:253])
            nome_pagador = line[324:354].strip() if len(line) > 354 else None
            data_credito = _cnab_parse_date(line[295:301]) if len(line) > 301 else None

            is_paid = codigo_ocorrencia in CNAB_PAID_CODES

            # Receivable: always create from detail (the boleto)
            receivables.append(Receivable(
                debtor_name=nome_pagador if nome_pagador else None,
                face_value=valor_titulo,
                due_date=vencimento,
                status="conciliated" if is_paid else "pending",
                source="cnab400",
            ))

            # Payment: only if paid
            if is_paid and valor_pago > 0:
                ref = nosso_numero or numero_documento
                payments.append(Payment(
                    payer_name=nome_pagador if nome_pagador else None,
                    amount=valor_pago,
                    date=data_credito or data_ocorrencia,
                    bank_reference=ref if ref else None,
                    source="cnab400",
                ))

        except Exception as e:
            errors.append({"row": line_num, "error": str(e)})

    return {"receivables": receivables, "payments": payments, "errors": errors}


def parse_cnab_content(content: bytes) -> dict:
    """Auto-detect CNAB 240 or 400 and parse."""
    fmt = _detect_cnab_format(content)
    if fmt == "240":
        return parse_cnab240_content(content)
    elif fmt == "400":
        return parse_cnab400_content(content)
    raise ValueError("Arquivo CNAB não reconhecido. Verifique se é um retorno CNAB 240 ou 400.")


def parse_file(content: bytes, filename: str) -> dict:
    """Route to the correct parser based on file extension."""
    if filename.endswith(".csv"):
        return parse_csv_content(content)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        return parse_xlsx_content(content)
    elif filename.endswith(".ofx"):
        return parse_ofx_content(content)
    elif filename.endswith(".ret") or filename.endswith(".rem") or filename.endswith(".cnab"):
        return parse_cnab_content(content)
    else:
        # Try CNAB auto-detect for extensionless or .txt files
        cnab_fmt = _detect_cnab_format(content)
        if cnab_fmt:
            return parse_cnab_content(content)
        raise ValueError(f"Formato não suportado: {filename}. Use CSV, XLSX, OFX ou CNAB.")
